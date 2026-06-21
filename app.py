"""
🍎 과일 위탁판매 발주 대시보드
쿠팡 API 기반 실시간 도매 발주 관리 시스템
"""

import streamlit as st
import hashlib
import hmac
import datetime
import json
import http.client
import requests
import random
import openpyxl
import pandas as pd
import re
from collections import defaultdict
from urllib.parse import urlencode

# ══════════════════════════════════════════
#  페이지 설정 (가장 먼저 실행)
# ══════════════════════════════════════════
st.set_page_config(
    page_title="🍎 과일 발주 대시보드",
    page_icon="🍎",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ══════════════════════════════════════════
#  전역 CSS (모바일 최적화)
# ══════════════════════════════════════════
st.markdown(
    """
<style>
/* ── 모바일 반응형 ── */
@media (max-width: 768px) {
    .main .block-container { padding: 0.8rem 0.4rem; }
    .stButton > button { font-size: 13px; padding: 6px 10px; }
    h1 { font-size: 22px !important; }
}

/* ── 공통 카드 ── */
.dash-header {
    background: linear-gradient(135deg, #FF6B35 0%, #F7931E 55%, #FFCD3C 100%);
    padding: 20px 24px;
    border-radius: 16px;
    margin-bottom: 20px;
    color: white;
}
.dash-header h1 { margin: 0; font-size: 26px; }
.dash-header p  { margin: 4px 0 0; opacity: .88; font-size: 14px; }

/* ── 도매사 탭 헤더 ── */
.tab-header {
    background: #f8fafc;
    border-left: 5px solid #FF6B35;
    border-radius: 10px;
    padding: 12px 16px;
    margin-bottom: 14px;
}
.tab-header h4 { margin: 0; color: #FF6B35; font-size: 17px; }
.tab-header span { color: #64748b; font-size: 13px; }

/* ── 이익 색상 ── */
.profit-pos { color: #16a34a; font-weight: 700; }
.profit-neg { color: #dc2626; font-weight: 700; }

/* ── 로그인 박스 ── */
.login-hint {
    text-align: center;
    color: #94a3b8;
    font-size: 12px;
    margin-top: 10px;
}

/* ── 복사 코드블록 ── */
.copy-block {
    background: #f1f5f9;
    border: 1px solid #cbd5e1;
    border-radius: 8px;
    padding: 14px 16px;
    font-family: 'Courier New', monospace;
    font-size: 14px;
    line-height: 2;
    white-space: pre-wrap;
    word-break: break-all;
}

/* ── 사이드바 헤더 ── */
.sb-header {
    background: linear-gradient(135deg, #FF6B35, #FFCD3C);
    padding: 10px 14px;
    border-radius: 10px;
    color: white;
    text-align: center;
    margin-bottom: 14px;
}
.sb-header h3 { margin: 0; font-size: 16px; }

/* Streamlit 기본 여백 숨김 */
#MainMenu { visibility: hidden; }
footer    { visibility: hidden; }
</style>
""",
    unsafe_allow_html=True,
)


# ══════════════════════════════════════════
#  세션 상태 초기화
# ══════════════════════════════════════════
def _secret(key: str, default: str = "") -> str:
    """secrets.toml 값 안전하게 읽기 (없어도 에러 없이 default 반환)."""
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


def _init():
    defaults = {
        "logged_in": False,
        "product_master": [],       # 상품 마스터 리스트
        "orders": [],               # 수집된 주문 리스트
        # API 키/시트 URL은 secrets.toml에 저장해두면 새로고침·재접속해도 유지됨
        "api_access_key": _secret("coupang_access_key"),
        "api_secret_key": _secret("coupang_secret_key"),
        "api_vendor_id":  _secret("coupang_vendor_id"),
        "dialog_order": None,       # 팝업에 표시할 주문
        "gsheet_url": _secret("gsheet_url"),  # 결제 기록/상품마스터/주문캐시용 구글 시트 URL
        "_recent_payments": [],     # 최근 결제 기록 캐시
        "_synced_from_sheet": False,  # 앱 시작 시 구글시트에서 1회 자동 불러오기 여부
        "_pm_edit_idx": None,       # 현재 수정 중인 상품 마스터 인덱스
        "_pm_form_key": 0,          # 상품 추가 폼 리셋 카운터 (등록 성공 시에만 증가)
        "naver_client_id":     _secret("naver_client_id"),
        "naver_client_secret": _secret("naver_client_secret"),
        "_discovery_rows": [],      # 신상품 발굴 분석 결과 캐시
        "_vendor_items_cache": [],  # 원가비교 페이지 업로드 데이터 캐시 (페이지 이동해도 유지)
        "_vendor_compare_summary": {},  # 원가비교에서 추출된 품목별 동시 취급 업체 수
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init()


# ══════════════════════════════════════════
#  로그인 시스템
# ══════════════════════════════════════════
def _get_password() -> str:
    """secrets.toml 우선, 없으면 기본값."""
    try:
        return st.secrets["app_password"]
    except Exception:
        return "fruit2024!"


def _auth_token() -> str:
    """현재 비밀번호로부터 만들어지는 고정 토큰. URL에 담아두면 새로고침(F5)해도 로그인 유지됨."""
    return hashlib.sha256(("fruit-dashboard-auth-" + _get_password()).encode("utf-8")).hexdigest()[:24]


def _check_url_auth():
    """URL의 인증 토큰이 유효하면 세션을 자동으로 로그인 상태로 만듦 (F5 새로고침 대응)."""
    try:
        token = st.query_params.get("k")
    except Exception:
        token = None
    if token and token == _auth_token():
        st.session_state.logged_in = True


def login_page():
    """로그인 UI만 렌더링하고 False를 반환."""
    st.markdown(
        """
        <div style="text-align:center; padding: 48px 16px 24px;">
            <div style="font-size:64px;">🍎</div>
            <h2 style="margin:8px 0 4px;">과일 발주 대시보드</h2>
            <p style="color:#64748b;">쿠팡 위탁판매 실시간 발주 관리 시스템</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns([1, 1.2, 1])
    with c2:
        with st.container(border=True):
            st.markdown("#### 🔐 로그인")
            pw = st.text_input("비밀번호", type="password", placeholder="비밀번호 입력")
            if st.button("로그인", use_container_width=True, type="primary"):
                if pw == _get_password():
                    st.session_state.logged_in = True
                    try:
                        st.query_params["k"] = _auth_token()
                    except Exception:
                        pass
                    st.rerun()
                else:
                    st.error("❌ 비밀번호가 올바르지 않습니다.")
            # 비밀번호를 secrets.toml에 아직 설정 안 했을 때만 운영자에게 보이는 경고
            # (화면에 실제 비밀번호 값은 절대 노출하지 않음)
            try:
                _has_custom_pw = "app_password" in st.secrets
            except Exception:
                _has_custom_pw = False
            if not _has_custom_pw:
                st.caption(
                    "⚠️ 기본 비밀번호가 사용 중입니다. "
                    ".streamlit/secrets.toml에 app_password를 설정해 변경해주세요."
                )
    return False


# ══════════════════════════════════════════
#  쿠팡 API 유틸
# ══════════════════════════════════════════
def _make_query_string(params: dict) -> str:
    """
    쿠팡 API 서명용 쿼리 스트링 생성.
    - 키 알파벳 오름차순 정렬
    - RFC 3986 percent-encoding (quote, safe='')
    - urlencode 미사용 → '+' 대신 '%20' 보장, 쿠팡 서버와 동일한 인코딩
    """
    from urllib.parse import quote as _quote
    pairs = sorted(params.items(), key=lambda x: x[0])
    return "&".join(f"{_quote(str(k), safe='')}={_quote(str(v), safe='')}" for k, v in pairs)


def _generate_auth_header(method: str, path: str, query_string: str,
                           access_key: str, secret_key: str) -> tuple[str, str]:
    """
    쿠팡 HMAC-SHA256 Authorization 헤더 생성.
    반환: (auth_header, signed_datetime)
    서명 메시지 형식 (쿠팡 공식 Python 예제 기준): {datetime}{METHOD}{path}{queryString}
    ※ 구분자 없이 그냥 이어붙임 (줄바꿈 X, '?' 도 넣지 않음)
    """
    dt = datetime.datetime.utcnow().strftime("%y%m%dT%H%M%SZ")
    message = f"{dt}{method}{path}{query_string}"
    signature = hmac.new(
        secret_key.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    auth = (
        f"CEA algorithm=HmacSHA256, access-key={access_key}, "
        f"signed-date={dt}, signature={signature}"
    )
    return auth, dt


def _official_hmac(method: str, full_url: str, access_key: str, secret_key: str) -> tuple[str, str, str]:
    """
    쿠팡 공식 SDK 패턴을 그대로 따른 HMAC 생성.
    full_url 에서 path 와 query_string 을 추출해 서명.
    반환: (auth_header, datetime_str, debug_message)
    서명 메시지 형식 (쿠팡 공식 Python 예제 기준): {datetime}{METHOD}{path}{queryString}
    ※ 구분자 없이 그냥 이어붙임 (줄바꿈 X, '?' 도 넣지 않음)
    """
    stripped = full_url.replace("https://api-gateway.coupang.com", "")
    dt = datetime.datetime.utcnow().strftime("%y%m%dT%H%M%SZ")

    if "?" in stripped:
        path_part, qs = stripped.split("?", 1)
    else:
        path_part, qs = stripped, ""

    message = f"{dt}{method}{path_part}{qs}"
    signature = hmac.new(
        secret_key.encode("utf-8"),
        message.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    auth = (
        f"CEA algorithm=HmacSHA256, access-key={access_key}, "
        f"signed-date={dt}, signature={signature}"
    )
    return auth, dt, message   # message 는 디버그용


def get_server_ip() -> str:
    """Streamlit 서버의 실제 외부 IP 주소 조회."""
    try:
        resp = requests.get("https://api.ipify.org", timeout=5)
        return resp.text.strip()
    except Exception as e:
        return f"오류: {e}"


def fetch_coupang_orders(access_key: str, secret_key: str, vendor_id: str,
                         debug: bool = False):
    """
    쿠팡 오픈 API – 발주서 리스트 조회 (오늘 날짜 기준).
    - http.client 사용: requests의 URL 재인코딩 문제를 완전히 우회
    - 쿼리 파라미터 값의 콜론(:)을 RFC 3986 기준 %3A로 인코딩
    - 파라미터 알파벳 오름차순 정렬: createdAtFrom < createdAtTo < pageSize < status
    debug=True 이면 서명 세부 정보를 st.expander 에 표시.
    """
    HOST   = "api-gateway.coupang.com"
    BASE   = f"https://{HOST}"
    path   = f"/v2/providers/openapi/apis/api/v4/vendors/{vendor_id}/ordersheets"
    method = "GET"

    today = datetime.date.today().isoformat()
    # RFC 3986: 콜론(:)을 %3A로 인코딩 (쿠팡 서버 정규화 기준에 맞춤)
    # 알파벳 오름차순: createdAtFrom < createdAtTo < pageSize < status
    qs = (
        f"createdAtFrom={today}T00%3A00%3A00"
        f"&createdAtTo={today}T23%3A59%3A59"
        f"&pageSize=50"
        f"&status=ACCEPT"
    )
    full_url = f"{BASE}{path}?{qs}"

    auth, dt, dbg_msg = _official_hmac(method, full_url, access_key, secret_key)

    if debug:
        with st.expander("🔍 API 서명 디버그 정보 (401 발생 시 확인용)", expanded=True):
            st.code(f"[Datetime]  {dt}\n[Method]    {method}\n[Path]      {path}\n[Query]     {qs}", language=None)
            st.markdown("**서명 메시지 (구분자 없이 이어붙임):**")
            st.code(dbg_msg, language=None)
            st.markdown("**요청 URL:**")
            st.code(full_url, language=None)
            st.markdown("**Authorization 헤더:**")
            st.code(auth, language=None)

    headers_dict = {
        "Authorization": auth,
        "Content-Type":  "application/json;charset=UTF-8",
        # 서버가 gzip으로 응답할 수 있어 명시적으로 비압축 요청
        # (그래도 gzip이 올 수 있어 아래에서 Content-Encoding을 확인해 안전하게 처리)
        "Accept-Encoding": "identity",
    }

    # http.client 사용 → URL 재인코딩 없이 정확히 전송
    conn = http.client.HTTPSConnection(HOST, timeout=10)
    conn.request("GET", f"{path}?{qs}", headers=headers_dict)
    resp = conn.getresponse()
    raw_body = resp.read()
    conn.close()

    # 응답이 gzip으로 압축되어 온 경우 압축 해제
    content_encoding = resp.getheader("Content-Encoding", "").lower()
    if content_encoding == "gzip" or raw_body[:2] == b"\x1f\x8b":
        import gzip
        raw_body = gzip.decompress(raw_body)

    resp_body = raw_body.decode("utf-8")

    if resp.status != 200:
        raise Exception(f"HTTP 오류 {resp.status}: {resp_body[:400]}")

    data = json.loads(resp_body)
    raw_sheets = data.get("data", {}).get("orderSheets", [])
    return _parse_sheets(raw_sheets)


def _parse_sheets(sheets: list) -> list:
    """쿠팡 API 응답 → 내부 주문 딕셔너리 리스트."""
    orders = []
    for sheet in sheets:
        receiver = sheet.get("receiver", {})
        orderer  = sheet.get("orderer", {})
        addr = (receiver.get("addr1", "") + " " + receiver.get("addr2", "")).strip()
        for item in sheet.get("items", []):
            orders.append({
                "order_id":        str(sheet.get("orderId", "")),
                "ordered_at":      sheet.get("orderedAt", ""),
                "option_id":       str(item.get("vendorItemId", "")),   # 쿠팡등록번호
                "product_name":    item.get("vendorItemName", ""),
                "quantity":        int(item.get("shippingCount", 1)),
                "sale_price":      int(item.get("salePrice", 0)),
                "orderer_name":    orderer.get("name", ""),
                "phone":           orderer.get("safeNumber", ""),
                "address":         addr,
                "delivery_message": receiver.get("parcelPrintMessage", ""),
                "status":          sheet.get("status", "ACCEPT"),
            })
    return orders


# ══════════════════════════════════════════
#  엑셀(발주서) 업로드 파싱
#  쿠팡 Wing → 주문/배송 → 발주서 조회 → 엑셀 다운로드 (DeliveryList_*.xlsx)
# ══════════════════════════════════════════
def parse_delivery_excel(uploaded_file) -> list:
    """
    쿠팡 Wing 발주서 조회 다운로드 엑셀(DeliveryList)을 내부 주문 형식으로 변환.
    헤더 컬럼명을 기준으로 매칭하므로, 컬럼 순서가 바뀌어도 안전하게 동작.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_name = "Delivery" if "Delivery" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    header = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    col_idx = {name: i for i, name in enumerate(header) if name}

    required = ["주문번호", "등록상품명", "옵션ID", "구매수(수량)", "옵션판매가(판매단가)"]
    missing = [r for r in required if r not in col_idx]
    if missing:
        raise ValueError(
            f"필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}\n"
            "쿠팡 Wing '주문/배송 → 발주서 조회' 메뉴에서 다운로드한 엑셀(DeliveryList)이 맞는지 확인해주세요."
        )

    def cell(row, name, default=""):
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        return val if val not in (None, "") else default

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not cell(row, "주문번호"):
            continue

        try:
            qty = int(float(cell(row, "구매수(수량)", 1) or 1))
        except (ValueError, TypeError):
            qty = 1
        try:
            price = int(float(cell(row, "옵션판매가(판매단가)", 0) or 0))
        except (ValueError, TypeError):
            price = 0

        orderer_name = cell(row, "수취인이름") or cell(row, "구매자")
        phone        = cell(row, "수취인전화번호") or cell(row, "구매자전화번호")

        base_name   = str(cell(row, "등록상품명"))
        option_name = str(cell(row, "등록옵션명"))
        # 등록상품명만으로는 무게/수량 옵션(0.5kg vs 1kg 등)이 구분 안 되므로 옵션명을 합쳐서 표시
        full_name = f"{base_name} ({option_name})" if option_name else base_name

        orders.append({
            "order_id":        str(cell(row, "주문번호")),
            "ordered_at":      str(cell(row, "주문일")),
            "option_id":       str(cell(row, "옵션ID")),
            "product_name":    full_name,
            "quantity":        qty,
            "sale_price":      price,
            "orderer_name":    str(orderer_name),
            "phone":           str(phone),
            "address":         str(cell(row, "수취인 주소")),
            "delivery_message": str(cell(row, "배송메세지")),
            "status":          "ACCEPT",
        })
    return orders


# ══════════════════════════════════════════
#  신상품 발굴 — 도매처 단가표 정밀 분석
#  (품목명 / 중량 / 등급 / 개수구간을 각각 분리 추출해서,
#   "1kg 대과"와 "1kg 중과"처럼 서로 다른 규격을 절대 같은 행으로 합치지 않음)
# ══════════════════════════════════════════

# 중량 패턴: "1kg", "500g", "10kg" 등 (숫자+단위를 그대로 캡처)
_WEIGHT_PATTERN = re.compile(r'\d+(\.\d+)?\s?(kg|KG|Kg|g|G|ml|ML)\b')

# 곱하기/묶음 표기: "500g x 2", "500g*2", "2팩" 등 — 다팩 구성을 별도 태그로 보존
_MULTIPLIER_PATTERN = re.compile(r'[*xX×]\s?\d+\s?(팩|포|봉|개)?')

# 개당 낱개 중량: "개당 30g내외" 처럼 진짜 박스 중량(1kg)과 별개로, 알맹이 하나의 크기를 나타내는 표기
# (개당 중량이 클수록 같은 등급이어도 더 굵은 과실 — 등급을 못 잡을 때 대체 비교 지표로 사용)
_UNIT_WEIGHT_PATTERN = re.compile(r'개당\s?\d+(\.\d+)?\s?(g|kg|G|KG)')

# 밀리미터 사이즈코드: 블루베리 등 알 크기를 "14-16mm", "18 20mm"처럼 표기
_MM_SIZE_PATTERN = re.compile(r'(\d+)\s?-?\s?(\d*)\s?mm', re.IGNORECASE)

# 개수구간 패턴: "13-16과", "8-12과 내외", "18과", "13과내외", "1번과"/"1번"/"3-4번"(등급을 숫자로 매긴 표기),
# "2팩"/"3포"/"1봉"(낱개 포장단위) 등
# 한글은 단어경계(\b)가 인식 안 되는 경우가 많아 \b를 쓰지 않음 (예: "13과내외"에서 과/내 사이는 경계가 아님)
_COUNT_RANGE_PATTERN = re.compile(r'\d+\s?[-~]\s?\d+\s?(번과|번|과|개|미|입|구|수|팩|포|봉)')
_COUNT_SINGLE_PATTERN = re.compile(r'\d+\s?(번과|번|과|개|미|입|구|수|팩|포|봉)')

# 감귤류 영문 사이즈코드: "2S", "S-M", "3L", "S,M,L"(여러 사이즈 혼합) 등
_CITRUS_SIZE_PATTERN = re.compile(r'(?<![A-Za-z0-9])([23]?[SML])(?:[-,/][23]?[SML])*(?![A-Za-z0-9])')

# 포장형태 — 가격에 영향을 주는 정보라 지우지 않고 별도 태그로 보존
_PACKAGE_WORDS = ['선물박스', '부직포가방', '스티로폼박스', '보자기', '지함']

# 사이즈 등급 단어 — 2글자 이상은 부분일치로 매칭 (길이순 정렬 필수: "중대과"가 "중과"로 잘못 잘리는 것 방지)
# 품질 등급 체계: 하품 < 중하 < 중품 < 중상 < 상품 < 특품 < 특상  (도매처마다 표현이 조금씩 다름)
_SIZE_GRADE_WORDS = ['특대과', '중대과', '중소과', '소중과', '왕왕특', '로얄과', '로열과', '특상품', '중상품', '중하품',
                     '왕미니', '꼬마과', '중상', '중하', '특상',
                     '대과', '중과', '소과', '왕특', '특품', '상품', '중품', '하품', '특대',
                     '로얄', '로열']

# 사이즈와 무관한 '타입/형태' 태그 — 등급과 별도로 함께 표시 (둘 다 동시에 있을 수 있음, 예: "소과 + 가정용")
_TYPE_TAGS = ['실속', '못난이', '정품', '가정용', '혼합', '랜덤', 'A급']

# 1글자 등급(하/중/상/특/대/소)은 공백으로 독립된 토큰일 때만 인정 (다른 단어 속 글자와 혼동 방지)
_GRADE_SINGLE_PATTERN = re.compile(r'(?:^|\s)([하중상특대소])(?:\s|$)')

# 품목명에서만 제거할 마케팅성 잡음 (수량/중량/등급은 절대 건드리지 않음)
_NOISE_WORDS = ['신규', 'NEW', 'New', 'new', '특가', '핫딜', '한정', '프리미엄', '단독', '이벤트',
                '순차출고', '순차 출고', '전후', '조기출고', '예약', '사전예약', '선물용']

_LOCATION_STOPWORDS = ['경북', '경남', '전남', '전북', '충남', '충북', '강원', '제주', '의성', '논산', '해남',
                       '청도', '나주', '영동', '통영', '목포', '여수', '산지', '직송', '박스포함', '박스',
                       '포장', '무료', '친환경', '무농약', '국내산', '수출용', '명품', '당도보장', '당도',
                       '세척', '비세척', '급냉', '세트', '선물세트', '증정', '1차', '2차', '내외', '이상', '미만']


def _strip_noise(text: str) -> str:
    """이모지·대괄호/별표 마케팅 문구·날짜/순차출고 같은 잡음 제거 (수량·중량·등급은 보존)."""
    tmp = text
    tmp = re.sub(r'[\U0001F300-\U0001FAFF\u2600-\u27BF★☆]', ' ', tmp)   # 이모지/별표
    tmp = re.sub(r'\[[^\]]*\]', ' ', tmp)                               # [마케팅용] 같은 대괄호 문구
    tmp = re.sub(r'\d{4}\.\s*\d{1,2}\.\s*\d{1,2}\..*$', ' ', tmp)       # 등록일시 꼬리표
    tmp = re.sub(r'\d{1,2}\s*월\s*\d{1,2}\s*일\s*~?', ' ', tmp)         # "6월 23일~" 같은 출고일 안내
    tmp = re.sub(r'\d{1,2}\s*/\s*\d{1,2}\s*', ' ', tmp)                 # "6/29" 같은 날짜 표기
    for w in _NOISE_WORDS:
        tmp = tmp.replace(w, ' ')
    return tmp


def parse_product_variant(full_name: str) -> dict:
    """
    상품명(+옵션명) 원문에서 품목명/중량/등급/개수구간/사이즈코드/포장형태를 각각 분리 추출.
    추출 순서대로 텍스트에서 잘라내며 진행하므로(겹침 방지), 추출 순서가 중요함:
    포장형태 → 곱하기묶음 → 개당중량 → 중량(kg/g) → mm사이즈 → 감귤사이즈코드 → 개수구간 → 등급
    """
    raw = str(full_name)

    # 포장형태 (선물박스/부직포가방 등) — 대괄호 안에 있어도 잡아야 하므로 노이즈 제거 전에 먼저 확인
    package = next((w for w in _PACKAGE_WORDS if w in raw), "")

    work = _strip_noise(raw)
    if package:
        work = work.replace(package, ' ')

    # 곱하기/묶음 표기 (예: "500g x 2", "500g*2", "*2팩") — 다팩 구성 별도 태그로 보존
    mult_m = _MULTIPLIER_PATTERN.search(work)
    multiplier = re.sub(r'\s+', '', mult_m.group(0)) if mult_m else ""
    if mult_m:
        work = work[:mult_m.start()] + ' ' + work[mult_m.end():]

    # 개당 낱개 중량 (예: "개당 30g내외") — 진짜 박스 중량(kg)보다 먼저 추출해야 "30g"이 중량으로 오인식 안 됨
    unit_w_m = _UNIT_WEIGHT_PATTERN.search(work)
    unit_weight = unit_w_m.group(0).replace(' ', '') if unit_w_m else ""
    if unit_w_m:
        work = work[:unit_w_m.start()] + ' ' + work[unit_w_m.end():]

    weight_m = _WEIGHT_PATTERN.search(work)
    weight = weight_m.group(0).replace(' ', '') if weight_m else ""
    if weight_m:
        work = work[:weight_m.start()] + ' ' + work[weight_m.end():]

    # mm 사이즈코드 (블루베리 등 알 크기 표기, 예: "18-20mm") — kg/L 단위와 안 겹치므로 중량 다음에 추출
    mm_m = _MM_SIZE_PATTERN.search(work)
    if mm_m:
        mm_size = f"{mm_m.group(1)}-{mm_m.group(2)}mm" if mm_m.group(2) else f"{mm_m.group(1)}mm"
        work = work[:mm_m.start()] + ' ' + work[mm_m.end():]
    else:
        mm_size = ""

    size_m = _CITRUS_SIZE_PATTERN.search(work)
    citrus_size = size_m.group(0).replace(' ', '') if size_m else ""
    if size_m:
        work = work[:size_m.start()] + ' ' + work[size_m.end():]

    count_m = _COUNT_RANGE_PATTERN.search(work) or _COUNT_SINGLE_PATTERN.search(work)
    count = count_m.group(0).replace(' ', '') if count_m else ""
    if count_m:
        work = work[:count_m.start()] + ' ' + work[count_m.end():]

    grade = ""
    for g in sorted(_SIZE_GRADE_WORDS, key=len, reverse=True):
        if g in work:
            grade = g
            break
    if not grade:
        single_m = _GRADE_SINGLE_PATTERN.search(work)
        if single_m:
            grade = single_m.group(1)
    if grade:
        if len(grade) == 1:
            work = re.sub(rf'(?:^|\s){re.escape(grade)}(?:\s|$)', ' ', work)
        else:
            work = work.replace(grade, ' ')

    # 사이즈 등급과는 별개로 '가정용/혼합/못난이' 같은 타입 태그도 따로 추출 (둘 다 동시에 있을 수 있음)
    type_tag = ""
    for t in sorted(_TYPE_TAGS, key=len, reverse=True):
        if t in work:
            type_tag = t
            work = work.replace(t, ' ')
            break

    base = re.sub(r'[()\[\]{}~/_,*xX×]', ' ', work)
    for w in _LOCATION_STOPWORDS:
        base = base.replace(w, ' ')
    base = re.sub(r'\s+', ' ', base).strip()

    # 비교 표시용 규격 라벨: "1kg 소과" / "S-M" / "1kg 중 (개당30g)" / "18-20mm" / "10kg 혼합 📦선물박스" 등
    spec_parts = [p for p in [weight, grade, type_tag, citrus_size, mm_size] if p]
    if unit_weight:
        spec_parts.append(f"({unit_weight})")
    if count and count != grade:
        spec_parts.append(f"({count})")
    if multiplier:
        spec_parts.append(multiplier)
    if package:
        spec_parts.append(f"📦{package}")
    spec_label = " ".join(spec_parts) if spec_parts else "규격미상"

    return {
        "base": base, "weight": weight, "grade": grade, "type_tag": type_tag, "count": count,
        "citrus_size": citrus_size, "unit_weight": unit_weight, "package": package,
        "spec_label": spec_label,
    }


def parse_vendor_excel(file_bytes, vendor_name: str, name_col, option_col, price_col, sheet_name=0) -> list:
    """업로드된 도매처 단가표 엑셀에서 (도매처, 원본상품명, 원가) 목록을 추출."""
    df = pd.read_excel(file_bytes, sheet_name=sheet_name)
    items = []
    for _, row in df.iterrows():
        parts = []
        if name_col and pd.notna(row.get(name_col)):
            parts.append(str(row[name_col]).strip())
        if option_col and option_col != "(없음)" and pd.notna(row.get(option_col)):
            parts.append(str(row[option_col]).strip())
        full_name = " ".join(p for p in parts if p)
        price = row.get(price_col)
        if not full_name or pd.isna(price):
            continue
        try:
            price = float(re.sub(r'[^\d.]', '', str(price)))
        except ValueError:
            continue
        if price <= 0:
            continue
        items.append({"공급사": vendor_name, "원본상품명": full_name, "원가": price})
    return items


def group_vendor_items(all_items: list) -> dict:
    """
    품목명(base) 기준 1차 그룹화 → 그 안에서 (중량, 등급, 개수구간)이 정확히 같은 것끼리만 2차로 묶어 비교.
    같은 '사과'라도 1kg/2kg, 대과/중과/소과는 서로 다른 행으로 분리된다.
    반환: { base명: { spec_label: {공급사: 최저원가, "_원본예시": str} } }
    """
    tree: dict = defaultdict(lambda: defaultdict(dict))
    for it in all_items:
        parsed = parse_product_variant(it["원본상품명"])
        base = parsed["base"]
        if len(base) < 2:
            continue
        spec = parsed["spec_label"]
        vendor = it["공급사"]
        bucket = tree[base][spec]
        if vendor not in bucket or it["원가"] < bucket[vendor]:
            bucket[vendor] = it["원가"]
        tree[base][spec]["_원본예시"] = it["원본상품명"]
    return tree


# ══════════════════════════════════════════
#  테스트 가상 주문 생성
_NAMES    = ["김철수", "이영희", "박민준", "최수연", "정도현", "강지은", "윤성호", "임나영"]
_PHONES   = ["010-1234-5678", "010-9876-5432", "010-5555-7777", "010-3333-4444"]
_ADDRS    = [
    "서울특별시 강남구 테헤란로 123 삼성아파트 101동 202호",
    "경기도 성남시 분당구 판교로 456 카카오빌딩 7층",
    "부산광역시 해운대구 해운대로 789 마린시티 30층",
    "인천광역시 연수구 송도과학로 90 아이파크 201동 1502호",
    "대구광역시 수성구 동대구로 100 범어아이파크 B동 304호",
]
_MSGS     = ["부재시 경비실에 맡겨주세요", "문 앞에 놓아주세요", "오전 배송 부탁드립니다", "조심히 다뤄주세요", ""]

_DEFAULT_PRODUCTS = [
    {"option_id": "TEST001", "product_name": "제주 천혜향 5kg",    "cost_price": 18000, "sale_price": 32000, "fee_rate": 10.8, "wholesale_name": "제주과일도매",   "wholesale_url": "https://naver.com"},
    {"option_id": "TEST002", "product_name": "충주 사과 10kg",     "cost_price": 22000, "sale_price": 38000, "fee_rate": 10.8, "wholesale_name": "충주사과마트",   "wholesale_url": ""},
    {"option_id": "TEST003", "product_name": "나주 배 선물세트 3개", "cost_price": 15000, "sale_price": 28000, "fee_rate": 10.8, "wholesale_name": "나주배농원",     "wholesale_url": "https://example.com"},
    {"option_id": "TEST004", "product_name": "샤인머스캣 2kg",      "cost_price": 25000, "sale_price": 45000, "fee_rate": 12.5, "wholesale_name": "제주과일도매",   "wholesale_url": "https://naver.com"},
]


def generate_test_orders() -> list:
    """상품 마스터(또는 기본 상품)를 바탕으로 가상 주문 10건 생성."""
    pool = st.session_state.product_master if st.session_state.product_master else _DEFAULT_PRODUCTS
    orders = []
    for i in range(10):
        prod = random.choice(pool)
        orders.append({
            "order_id":         f"TEST-{i+1:04d}",
            "ordered_at":       datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "option_id":        prod["option_id"],
            "product_name":     prod["product_name"],
            "quantity":         random.randint(1, 3),
            "sale_price":       prod.get("sale_price", 30000),
            "orderer_name":     random.choice(_NAMES),
            "phone":            random.choice(_PHONES),
            "address":          random.choice(_ADDRS),
            "delivery_message": random.choice(_MSGS),
            "status":           "ACCEPT",
        })
    return orders


# ══════════════════════════════════════════
#  수익 계산
# ══════════════════════════════════════════
def calc_profit(order: dict, master: dict) -> dict:
    qty        = order["quantity"]
    sale_price = order["sale_price"]
    cost_price = master.get("cost_price", 0)
    fee_rate   = master.get("fee_rate", 10.8) / 100

    revenue = sale_price * qty
    cost    = cost_price * qty
    fee     = revenue * fee_rate
    profit  = revenue - cost - fee
    margin  = (profit / revenue * 100) if revenue > 0 else 0

    return {"revenue": revenue, "cost": cost, "fee": fee, "profit": profit, "margin": margin}


# ══════════════════════════════════════════
#  발주 정보 팝업 (st.dialog)
# ══════════════════════════════════════════
@st.dialog("📋 발주 정보", width="large")
def order_dialog(order: dict, master: dict):
    product_name   = master.get("product_name", order["product_name"])
    wholesale_name = master.get("wholesale_name", "-")
    wholesale_url  = master.get("wholesale_url", "")

    st.markdown(f"### 🛒 {product_name}")
    st.caption(f"도매사: **{wholesale_name}** | 주문 ID: `{order['order_id']}`")

    if wholesale_url:
        st.link_button("🚀 도매 사이트 바로가기", wholesale_url, type="primary")

    st.divider()

    # 필드별 코드블록
    fields = [
        ("👤 주문자 이름",           order["orderer_name"]),
        ("📞 연락처",                order["phone"]),
        ("📍 배송 주소",             order["address"]),
        ("💬 배송 메시지(요청사항)", order["delivery_message"] or "(없음)"),
        ("📦 수량",                  f"{order['quantity']}개"),
    ]
    for label, val in fields:
        st.markdown(f"**{label}**")
        st.code(val, language=None)

    st.divider()
    st.markdown("#### 📋 도매 발주용 통합 복사")
    copy_text = (
        f"주문자: {order['orderer_name']}\n"
        f"연락처: {order['phone']}\n"
        f"주소: {order['address']}\n"
        f"배송메시지: {order['delivery_message'] or '없음'}\n"
        f"상품: {product_name} × {order['quantity']}개"
    )
    st.code(copy_text, language=None)


# ══════════════════════════════════════════
#  결제 기록(세금 증빙용) – 구글 시트 연동
# ══════════════════════════════════════════
def _get_gsheet_client():
    """구글 시트 연결 클라이언트 생성. secrets.toml의 [gcp_service_account] 섹션 필요."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise RuntimeError(
            "gspread / google-auth 패키지가 설치되어 있지 않습니다. "
            "requirements.txt에 'gspread'와 'google-auth'를 추가하고 재배포해주세요."
        )
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError(
            "구글 서비스 계정 정보가 없습니다. .streamlit/secrets.toml에 "
            "[gcp_service_account] 섹션을 추가해주세요. (사이드바 하단 '연동 설정 방법' 참고)"
        )
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=scopes
    )
    return gspread.authorize(creds)


def _get_or_create_worksheet(title: str, headers: list):
    """제목의 워크시트를 가져오거나, 없으면 헤더와 함께 새로 생성."""
    if not st.session_state.gsheet_url:
        raise RuntimeError("사이드바에서 구글 시트 URL을 먼저 입력해주세요.")
    client = _get_gsheet_client()
    sh = client.open_by_url(st.session_state.gsheet_url)
    try:
        ws = sh.worksheet(title)
    except Exception:
        ws = sh.add_worksheet(title=title, rows=2000, cols=max(len(headers), 1))
        ws.append_row(headers)
    return ws


_PAYMENT_HEADERS = ["기록일시", "주문ID", "도매처", "결제일시", "결제수단", "은행/카드사", "금액", "메모"]


def _get_payment_worksheet():
    """결제 기록용 워크시트 객체 반환. 없으면 헤더와 함께 새로 생성."""
    return _get_or_create_worksheet("결제기록", _PAYMENT_HEADERS)


def append_payment_record(record: dict):
    """결제 기록 1건을 구글 시트에 추가."""
    ws = _get_payment_worksheet()
    ws.append_row([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        record.get("order_id", ""),
        record.get("wholesale_name", ""),
        record.get("pay_datetime", ""),
        record.get("pay_method", ""),
        record.get("bank_or_card", ""),
        record.get("amount", 0),
        record.get("memo", ""),
    ])


def fetch_recent_payment_records(limit: int = 10) -> list:
    """최근 결제 기록 N건을 최신순으로 반환."""
    ws = _get_payment_worksheet()
    rows = ws.get_all_records()
    return list(reversed(rows))[:limit]


# ── 상품 마스터 저장/불러오기 ──
_PRODUCT_HEADERS = ["옵션ID", "제품이름", "원가", "판매가", "수수료율", "도매처", "도매URL"]


def save_product_master_to_sheet():
    """현재 상품 마스터 전체를 구글 시트에 덮어쓰기 저장."""
    ws = _get_or_create_worksheet("상품마스터", _PRODUCT_HEADERS)
    ws.clear()
    ws.append_row(_PRODUCT_HEADERS)
    rows = [
        [p["option_id"], p["product_name"], p["cost_price"], p["sale_price"],
         p["fee_rate"], p["wholesale_name"], p["wholesale_url"]]
        for p in st.session_state.product_master
    ]
    if rows:
        ws.append_rows(rows)


def load_product_master_from_sheet() -> list:
    """구글 시트에 저장된 상품 마스터를 불러옴."""
    ws = _get_or_create_worksheet("상품마스터", _PRODUCT_HEADERS)
    records = ws.get_all_records()
    result = []
    for r in records:
        if not r.get("옵션ID"):
            continue
        result.append({
            "option_id":     str(r.get("옵션ID", "")),
            "product_name":  r.get("제품이름", ""),
            "cost_price":    int(float(r.get("원가", 0) or 0)),
            "sale_price":    int(float(r.get("판매가", 0) or 0)),
            "fee_rate":      float(r.get("수수료율", 0) or 0),
            "wholesale_name":r.get("도매처", ""),
            "wholesale_url": r.get("도매URL", ""),
        })
    return result


# ── 주문 데이터 캐시 저장/불러오기 (새로고침해도 다시 업로드 안 해도 됨) ──
_ORDER_HEADERS = [
    "order_id", "ordered_at", "option_id", "product_name", "quantity",
    "sale_price", "orderer_name", "phone", "address", "delivery_message", "status",
]


def save_orders_cache_to_sheet(orders: list):
    """현재 주문 목록 전체를 구글 시트에 덮어쓰기 저장 (마지막 동기화 스냅샷)."""
    ws = _get_or_create_worksheet("주문캐시", _ORDER_HEADERS)
    ws.clear()
    ws.append_row(_ORDER_HEADERS)
    rows = [[o.get(h, "") for h in _ORDER_HEADERS] for o in orders]
    if rows:
        ws.append_rows(rows)


def load_orders_cache_from_sheet() -> list:
    """구글 시트에 저장된 마지막 주문 스냅샷을 불러옴."""
    ws = _get_or_create_worksheet("주문캐시", _ORDER_HEADERS)
    records = ws.get_all_records()
    orders = []
    for r in records:
        if not r.get("order_id"):
            continue
        orders.append({
            "order_id":        str(r.get("order_id", "")),
            "ordered_at":      str(r.get("ordered_at", "")),
            "option_id":       str(r.get("option_id", "")),
            "product_name":    str(r.get("product_name", "")),
            "quantity":        int(float(r.get("quantity", 1) or 1)),
            "sale_price":      int(float(r.get("sale_price", 0) or 0)),
            "orderer_name":    str(r.get("orderer_name", "")),
            "phone":           str(r.get("phone", "")),
            "address":         str(r.get("address", "")),
            "delivery_message":str(r.get("delivery_message", "")),
            "status":          str(r.get("status", "ACCEPT")),
        })
    return orders


# ── 원가비교(도매처 단가) 캐시 저장/불러오기 — 구글시트 (세션이 끊겨도 영구 유지) ──
_VENDOR_ITEMS_HEADERS = ["공급사", "원본상품명", "원가"]


def save_vendor_items_to_sheet(all_items: list):
    """원가비교에서 파싱한 원본 항목 전체를 구글 시트에 덮어쓰기 저장."""
    ws = _get_or_create_worksheet("원가비교캐시", _VENDOR_ITEMS_HEADERS)
    ws.clear()
    ws.append_row(_VENDOR_ITEMS_HEADERS)
    rows = [[it["공급사"], it["원본상품명"], it["원가"]] for it in all_items]
    if rows:
        ws.append_rows(rows)


def load_vendor_items_from_sheet() -> list:
    """구글 시트에 저장된 원가비교 원본 항목을 불러옴."""
    ws = _get_or_create_worksheet("원가비교캐시", _VENDOR_ITEMS_HEADERS)
    records = ws.get_all_records()
    items = []
    for r in records:
        if not r.get("원본상품명"):
            continue
        try:
            price = float(r.get("원가", 0) or 0)
        except (ValueError, TypeError):
            price = 0
        items.append({
            "공급사": str(r.get("공급사", "")),
            "원본상품명": str(r.get("원본상품명", "")),
            "원가": price,
        })
    return items


# ══════════════════════════════════════════
#  사이드바 – 상품 마스터 관리
# ══════════════════════════════════════════
def _render_product_master_panel():
    if st.session_state.gsheet_url:
        sc1, sc2 = st.columns(2)
        if sc1.button("☁️ 시트에서 불러오기", use_container_width=True, key="btn_load_pm"):
            try:
                st.session_state.product_master = load_product_master_from_sheet()
                st.success(f"✅ {len(st.session_state.product_master)}개 상품 불러옴")
                st.rerun()
            except Exception as e:
                st.error(f"❌ {e}")
        if sc2.button("💾 시트에 저장", use_container_width=True, key="btn_save_pm"):
            try:
                save_product_master_to_sheet()
                st.success("✅ 저장 완료")
            except Exception as e:
                st.error(f"❌ {e}")
    else:
        st.caption("💡 사이드바 → 💰 결제 기록에서 구글시트를 연동하면, 여기 등록한 상품도 자동 저장되어 다음 접속 때 다시 입력 안 해도 됩니다.")

    # API 키 설정
    with st.expander("⚙️ 쿠팡 API 설정", expanded=False):
        st.session_state.api_access_key = st.text_input(
            "Access Key", value=st.session_state.api_access_key,
            type="password", key="inp_ak",
        )
        st.session_state.api_secret_key = st.text_input(
            "Secret Key", value=st.session_state.api_secret_key,
            type="password", key="inp_sk",
        )
        st.session_state.api_vendor_id = st.text_input(
            "Vendor ID", value=st.session_state.api_vendor_id, key="inp_vid",
        )

    # 기존 등록된 도매사 목록 (이름→URL) — 반복 입력 줄이기용 자동완성
    existing_wholesalers = {}
    for p in st.session_state.product_master:
        if p.get("wholesale_name"):
            existing_wholesalers[p["wholesale_name"]] = p.get("wholesale_url", "")

    # 상품 추가 폼
    with st.expander("➕ 상품 추가", expanded=True):
        default_wname, default_wurl = "", ""
        if existing_wholesalers:
            picked = st.selectbox(
                "🏪 기존 도매사에서 선택 (선택하면 아래 도매사 정보 자동 입력)",
                ["+ 새 도매사 직접 입력"] + sorted(existing_wholesalers.keys()),
                key="pm_wholesale_picker",
            )
            if picked != "+ 새 도매사 직접 입력":
                default_wname, default_wurl = picked, existing_wholesalers[picked]

        # form_key는 등록 '성공'했을 때만 증가시켜서, 검증 실패 시 입력값이 유지되게 함
        form_key = f"product_form_{st.session_state._pm_form_key}"
        with st.form(form_key):
            option_id    = st.text_input("🔑 쿠팡등록번호(옵션ID)*", placeholder="예: 7654321098")
            product_name = st.text_input("📦 제품이름*", placeholder="예: 제주 천혜향 5kg")
            cost_price   = st.number_input("💰 원가(도매가, 원)", min_value=0, value=0, step=500)
            sale_price   = st.number_input("🏷️ 쿠팡 판매가(원)", min_value=0, value=0, step=500)
            fee_rate     = st.number_input("📊 판매 수수료율(%)", min_value=0.0, max_value=100.0, value=10.8, step=0.1)
            wholesale_nm = st.text_input("🏪 도매사 이름*", value=default_wname, placeholder="예: 제주과일도매")
            wholesale_url= st.text_input("🔗 도매 주문 페이지 URL", value=default_wurl, placeholder="https://...")

            if st.form_submit_button("✅ 상품 등록", use_container_width=True, type="primary"):
                if not option_id or not product_name or not wholesale_nm:
                    st.error("* 표시 필드는 필수입니다. (입력하신 내용은 그대로 남아있어요)")
                elif option_id in [p["option_id"] for p in st.session_state.product_master]:
                    st.warning(f"옵션ID '{option_id}'는 이미 등록되어 있습니다. (입력하신 내용은 그대로 남아있어요)")
                else:
                    st.session_state.product_master.append({
                        "option_id":     option_id,
                        "product_name":  product_name,
                        "cost_price":    cost_price,
                        "sale_price":    sale_price,
                        "fee_rate":      fee_rate,
                        "wholesale_name":wholesale_nm,
                        "wholesale_url": wholesale_url,
                    })
                    try:
                        save_product_master_to_sheet()
                    except Exception:
                        pass  # 구글시트 미설정 시 조용히 무시 (세션 내에서는 계속 사용 가능)
                    st.session_state._pm_form_key += 1  # 성공했을 때만 폼 리셋
                    st.success(f"✅ '{product_name}' 등록 완료!")
                    st.rerun()

    # ── 도매처 이름 일괄 병합 (예: "최고"와 "최고집"이 사실 같은 곳일 때 한 번에 합치기) ──
    wholesale_names = sorted({
        p["wholesale_name"] for p in st.session_state.product_master if p.get("wholesale_name")
    })
    if len(wholesale_names) >= 2:
        with st.expander("🔀 도매처 이름 병합/정리", expanded=False):
            st.caption("같은 도매처인데 이름이 다르게 저장된 경우, 여기서 한 번에 합칠 수 있어요.")
            mc1, mc2 = st.columns(2)
            src_name = mc1.selectbox("이 이름을", wholesale_names, key="merge_src")
            dst_options = [n for n in wholesale_names if n != src_name]
            dst_name = mc2.selectbox("이 이름으로 합치기", dst_options, key="merge_dst") if dst_options else None

            affected = [p for p in st.session_state.product_master if p.get("wholesale_name") == src_name]
            st.caption(f"'{src_name}' 으로 등록된 상품 {len(affected)}개가 '{dst_name}' 으로 변경됩니다.")

            if dst_name and st.button(f"🔀 '{src_name}' → '{dst_name}' 으로 병합", use_container_width=True, type="primary"):
                for p in st.session_state.product_master:
                    if p.get("wholesale_name") == src_name:
                        p["wholesale_name"] = dst_name
                        # URL이 비어있는 상품은 병합 대상 도매처의 URL로 같이 채워줌
                        if not p.get("wholesale_url"):
                            ref_url = next(
                                (q["wholesale_url"] for q in st.session_state.product_master
                                 if q.get("wholesale_name") == dst_name and q.get("wholesale_url")),
                                "",
                            )
                            if ref_url:
                                p["wholesale_url"] = ref_url
                try:
                    save_product_master_to_sheet()
                except Exception:
                    pass
                st.success(f"✅ {len(affected)}개 상품을 '{dst_name}' 으로 병합했습니다!")
                st.rerun()

    # ── 상품 수정 폼 (목록에서 ✏️ 누르면 바로 이 자리, 목록 바로 위에 표시됨) ──
    edit_idx = st.session_state._pm_edit_idx
    if edit_idx is not None and 0 <= edit_idx < len(st.session_state.product_master):
        p = st.session_state.product_master[edit_idx]
        with st.container(border=True):
            st.markdown(f"##### ✏️ 상품 수정 — {p['product_name']}")
            with st.form(f"edit_form_{edit_idx}"):
                e_oid   = st.text_input("🔑 쿠팡등록번호(옵션ID)*", value=p["option_id"])
                e_pname = st.text_input("📦 제품이름*", value=p["product_name"])
                e_cost  = st.number_input("💰 원가(도매가, 원)", min_value=0, value=p["cost_price"], step=500)
                e_sale  = st.number_input("🏷️ 쿠팡 판매가(원)", min_value=0, value=p["sale_price"], step=500)
                e_fee   = st.number_input("📊 판매 수수료율(%)", min_value=0.0, max_value=100.0, value=p["fee_rate"], step=0.1)
                e_wname = st.text_input("🏪 도매사 이름*", value=p["wholesale_name"])
                e_wurl  = st.text_input("🔗 도매 주문 페이지 URL", value=p["wholesale_url"])

                ec1, ec2 = st.columns(2)
                save_clicked   = ec1.form_submit_button("💾 수정 저장", use_container_width=True, type="primary")
                cancel_clicked = ec2.form_submit_button("취소", use_container_width=True)

                if save_clicked:
                    if not e_oid or not e_pname or not e_wname:
                        st.error("* 표시 필드는 필수입니다.")
                    else:
                        st.session_state.product_master[edit_idx] = {
                            "option_id":     e_oid,
                            "product_name":  e_pname,
                            "cost_price":    e_cost,
                            "sale_price":    e_sale,
                            "fee_rate":      e_fee,
                            "wholesale_name":e_wname,
                            "wholesale_url": e_wurl,
                        }
                        try:
                            save_product_master_to_sheet()
                        except Exception:
                            pass
                        st.session_state._pm_edit_idx = None
                        st.success("✅ 수정 완료!")
                        st.rerun()
                if cancel_clicked:
                    st.session_state._pm_edit_idx = None
                    st.rerun()

    # 등록 상품 목록
    count = len(st.session_state.product_master)
    st.markdown(f"#### 📋 등록 상품 ({count}개)")

    if st.session_state.product_master:
        for idx, p in enumerate(st.session_state.product_master):
            with st.container(border=True):
                c1, c2 = st.columns([5, 1.4])
                with c1:
                    st.markdown(f"**{p['product_name']}**")
                    st.caption(f"ID: `{p['option_id']}` | {p['wholesale_name']}")
                    if p["sale_price"] > 0:
                        raw_margin = (p["sale_price"] - p["cost_price"]) / p["sale_price"] * 100 - p["fee_rate"]
                        color = "#16a34a" if raw_margin >= 0 else "#dc2626"
                        st.markdown(
                            f"<small>판매가: ₩{p['sale_price']:,} | "
                            f"마진: <span style='color:{color};font-weight:700'>{raw_margin:.1f}%</span></small>",
                            unsafe_allow_html=True,
                        )
                with c2:
                    ec1, ec2 = st.columns(2)
                    if ec1.button("✏️", key=f"edit_{idx}", help="수정"):
                        st.session_state._pm_edit_idx = idx
                        st.rerun()
                    if ec2.button("🗑️", key=f"del_{idx}", help="삭제"):
                        st.session_state.product_master.pop(idx)
                        if st.session_state._pm_edit_idx == idx:
                            st.session_state._pm_edit_idx = None
                        try:
                            save_product_master_to_sheet()
                        except Exception:
                            pass
                        st.rerun()
    else:
        st.info("등록된 상품이 없습니다.\n위 폼에서 추가해주세요.")


# ══════════════════════════════════════════
#  사이드바 – 결제 기록 (세금 증빙용)
# ══════════════════════════════════════════
def _render_payment_panel():
    st.caption("도매처 결제 정보(날짜/은행/카드/금액)를 구글시트에 자동 저장합니다. 사진 첨부 없이 텍스트만 기록합니다.")

    with st.expander("🔧 구글시트 연동 설정", expanded=not st.session_state.gsheet_url):
        st.session_state.gsheet_url = st.text_input(
            "구글 시트 URL", value=st.session_state.gsheet_url,
            placeholder="https://docs.google.com/spreadsheets/d/...",
            key="inp_gsheet_url",
        )
        st.caption("⚠️ 시트를 서비스 계정 이메일과 '편집자' 권한으로 공유해야 합니다. 맨 아래 '연동 설정 방법' 참고.")
        st.caption(
            "💡 여기 입력한 URL은 이번 접속에서만 유지됩니다. **새로고침해도 계속 유지**하려면 "
            "`.streamlit/secrets.toml`에 `gsheet_url = \"여기에 URL\"` 한 줄을 추가해주세요 "
            "(아래 '연동 설정 방법'에 같이 안내되어 있습니다)."
        )

    wholesale_options = sorted({
        p["wholesale_name"] for p in st.session_state.product_master if p.get("wholesale_name")
    })
    order_id_options = [""] + [o["order_id"] for o in st.session_state.orders]

    with st.expander("➕ 결제 기록 추가", expanded=True):
        with st.form("payment_form", clear_on_submit=True):
            order_id = st.selectbox("🔗 연결할 주문 ID (선택)", order_id_options)
            if wholesale_options:
                wholesale_name = st.selectbox("🏪 도매처", wholesale_options)
            else:
                wholesale_name = st.text_input("🏪 도매처", placeholder="예: 제주과일도매")
            pcol1, pcol2 = st.columns(2)
            pay_date = pcol1.date_input("📅 결제 날짜")
            pay_time = pcol2.time_input("🕒 결제 시각")
            pay_method = st.selectbox("💳 결제 수단", ["카드", "현금", "계좌이체"])
            bank_or_card = st.text_input("🏦 은행명 / 카드사", placeholder="예: 국민은행, 신한카드")
            amount = st.number_input("💰 금액(원)", min_value=0, value=0, step=1000)
            memo = st.text_input("📝 메모", placeholder="예: 6/19 주문 5건 일괄결제")

            if st.form_submit_button("✅ 결제 기록 저장", use_container_width=True, type="primary"):
                if not st.session_state.gsheet_url:
                    st.error("구글 시트 URL을 먼저 설정해주세요.")
                elif amount <= 0:
                    st.error("금액을 입력해주세요.")
                else:
                    try:
                        append_payment_record({
                            "order_id":      order_id,
                            "wholesale_name":wholesale_name,
                            "pay_datetime":  f"{pay_date} {pay_time.strftime('%H:%M')}",
                            "pay_method":    pay_method,
                            "bank_or_card":  bank_or_card,
                            "amount":        amount,
                            "memo":          memo,
                        })
                        st.success("✅ 결제 기록이 구글시트에 저장되었습니다!")
                    except Exception as e:
                        st.error(f"❌ {e}")

    st.divider()
    st.markdown("#### 🧾 최근 결제 기록")
    if st.session_state.gsheet_url:
        if st.button("🔄 최근 기록 불러오기", use_container_width=True):
            try:
                st.session_state._recent_payments = fetch_recent_payment_records(10)
            except Exception as e:
                st.error(f"❌ {e}")

        if st.session_state._recent_payments:
            for r in st.session_state._recent_payments:
                with st.container(border=True):
                    st.caption(f"{r.get('결제일시','')} · {r.get('도매처','')}")
                    amt = r.get("금액", 0)
                    try:
                        amt = int(amt)
                    except (ValueError, TypeError):
                        amt = 0
                    st.markdown(f"**₩{amt:,}** · {r.get('결제수단','')} ({r.get('은행/카드사','')})")
                    if r.get("메모"):
                        st.caption(f"📝 {r['메모']}")
        else:
            st.caption("아직 불러온 기록이 없습니다. 위 버튼을 눌러주세요.")
    else:
        st.info("구글 시트 URL을 먼저 설정해주세요.")

    with st.expander("📖 구글시트 연동 설정 방법 (최초 1회만)", expanded=False):
        st.markdown(
            "1. **구글 시트**에서 빈 시트 하나 새로 만들고 URL 복사\n"
            "2. **Google Cloud Console**(console.cloud.google.com) → 새 프로젝트 생성\n"
            "3. **API 라이브러리**에서 `Google Sheets API`, `Google Drive API` 활성화\n"
            "4. **사용자 인증정보 → 서비스 계정 만들기** → 생성 후 '키' 탭에서 JSON 키 다운로드\n"
            "5. 다운로드한 JSON 내용 + 시트 URL을 `.streamlit/secrets.toml`에 아래 형식으로 붙여넣기 "
            "(URL을 여기에 같이 넣어두면 **새로고침·재접속해도 자동으로 다시 연결**됩니다):\n"
        )
        st.code(
            'gsheet_url = "https://docs.google.com/spreadsheets/d/여기에_시트ID/edit"\n\n'
            '[gcp_service_account]\n'
            'type = "service_account"\n'
            'project_id = "..."\n'
            'private_key_id = "..."\n'
            'private_key = "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n"\n'
            'client_email = "...@....iam.gserviceaccount.com"\n'
            'client_id = "..."\n'
            'token_uri = "https://oauth2.googleapis.com/token"',
            language="toml",
        )
        st.markdown(
            "6. JSON 안의 `client_email` 값을 복사해서, **1번 시트를 그 이메일과 '편집자' 권한으로 공유**\n"
            "7. requirements.txt에 `gspread`, `google-auth` 추가 후 재배포\n"
            "8. 재배포하면 자동으로 연결되어 있을 거예요 (위 입력칸에 매번 URL 안 넣어도 됨)\n\n"
            "💡 Streamlit Community Cloud에 배포 중이라면, 로컬 파일 대신 "
            "**앱 설정(Settings) → Secrets** 메뉴에 위 내용을 그대로 붙여넣으시면 됩니다."
        )


def render_sidebar():
    with st.sidebar:
        st.markdown(
            '<div class="sb-header"><h3>🍊 발주 대시보드 관리</h3>'
            "<small>상품 마스터 & 결제 기록</small></div>",
            unsafe_allow_html=True,
        )

        nav = st.radio(
            "사이드바 카테고리",
            ["🍊 상품 마스터", "💰 결제 기록(세금)"],
            horizontal=True,
            label_visibility="collapsed",
            key="sidebar_nav",
        )
        st.divider()

        if nav == "🍊 상품 마스터":
            _render_product_master_panel()
        else:
            _render_payment_panel()

        # 로그아웃
        st.divider()
        if st.button("🚪 로그아웃", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.orders = []
            try:
                if "k" in st.query_params:
                    del st.query_params["k"]
            except Exception:
                pass
            st.rerun()


# ══════════════════════════════════════════
#  신상품 발굴 — 네이버 데이터랩(트렌드) + 네이버쇼핑(시장가) 분석
# ══════════════════════════════════════════
NAVER_DATALAB_URL = "https://openapi.naver.com/v1/datalab/search"
NAVER_SHOP_URL = "https://openapi.naver.com/v1/search/shop.json"

# 채널별 수수료 (필요시 화면에서 직접 조정 가능)
_CHANNEL_FEES_DEFAULT = {
    "쿠팡": 0.108,
    "네이버 스마트스토어": 0.0374 + 0.02,
    "G마켓": 0.13,
}


@st.cache_data(ttl=86400, show_spinner=False)
def fetch_naver_trend_batch(keywords: tuple, client_id: str, client_secret: str, time_unit: str = "month") -> dict:
    """네이버 데이터랩으로 키워드별 최근 2년 검색 트렌드 조회 (5개씩 묶어서 호출).
    time_unit: 'date'(일간) / 'week'(주간) / 'month'(월간)
    """
    end_date = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=2 * 365)
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
        "Content-Type": "application/json",
    }
    all_results, errors = [], []
    items = list(keywords)
    for i in range(0, len(items), 5):
        chunk = items[i:i + 5]
        keyword_groups = [{"groupName": name, "keywords": [name]} for name in chunk]
        body = {
            "startDate": start_date.strftime("%Y-%m-%d"),
            "endDate": end_date.strftime("%Y-%m-%d"),
            "timeUnit": time_unit,
            "keywordGroups": keyword_groups,
        }
        try:
            resp = requests.post(NAVER_DATALAB_URL, headers=headers, data=json.dumps(body), timeout=10)
            if resp.status_code != 200:
                errors.append(f"{resp.status_code} {resp.text[:150]}")
                continue
            all_results.extend(resp.json().get("results", []))
        except Exception as e:
            errors.append(str(e))
    return {"results": all_results, "errors": errors}


def circular_month_distance(m1: int, m2: int) -> int:
    d = abs(m1 - m2) % 12
    return min(d, 12 - d)


def analyze_seasonality(trend_data: dict, keyword: str) -> dict:
    """월별 검색량 평균을 내서 피크월/에버그린 여부/최근 모멘텀(상승·하락세)을 판정."""
    # 공백/대소문자 차이로 매칭 실패하는 걸 막기 위해 느슨하게 비교
    norm = lambda s: re.sub(r'\s+', '', str(s)).lower()
    results_by_title = {norm(r.get("title")): r for r in trend_data.get("results", [])}
    r = results_by_title.get(norm(keyword))
    data_points = r.get("data", []) if r else []

    if r is None:
        return {"ok": False, "peak_month": None, "is_evergreen": False, "momentum": 0.0,
                "debug": "⚠️ API 응답에서 이 키워드를 찾지 못함 (네이버 데이터랩 매칭 실패)"}
    if not data_points or len(data_points) < 6:
        return {"ok": False, "peak_month": None, "is_evergreen": False, "momentum": 0.0,
                "debug": "데이터 부족 (최근 검색량 데이터가 6개월 미만)"}

    monthly = defaultdict(list)
    for d in data_points:
        try:
            m = int(d["period"].split("-")[1])
        except (KeyError, IndexError, ValueError):
            continue
        monthly[m].append(d["ratio"])

    if not monthly:
        return {"ok": False, "peak_month": None, "is_evergreen": False, "momentum": 0.0, "debug": "월별 파싱 실패"}

    monthly_avg = {m: sum(v) / len(v) for m, v in monthly.items()}
    peak_month = max(monthly_avg, key=monthly_avg.get)

    vals = list(monthly_avg.values())
    mean_v = sum(vals) / len(vals)
    std_v = (sum((x - mean_v) ** 2 for x in vals) / len(vals)) ** 0.5
    cv = (std_v / mean_v) if mean_v > 0 else 0

    last_point = data_points[-1]
    current_ratio = last_point["ratio"]
    try:
        cur_month = int(last_point["period"].split("-")[1])
    except (KeyError, IndexError, ValueError):
        cur_month = peak_month
    baseline = monthly_avg.get(cur_month, mean_v)
    momentum = ((current_ratio - baseline) / baseline) if baseline > 0 else 0.0

    return {
        "ok": True,
        "peak_month": peak_month,
        "is_evergreen": cv < 0.25,   # 월별 편차가 작으면 '사계절 꾸준히 팔리는' 에버그린형으로 판정
        "momentum": momentum,
        "debug": f"피크 {peak_month}월, 변동계수 {cv:.2f}",
    }


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_naver_shopping(query: str, client_id: str, client_secret: str) -> dict:
    """네이버쇼핑 검색으로 시장 평균가·최저가·판매처 수 조회."""
    headers = {"X-Naver-Client-Id": client_id, "X-Naver-Client-Secret": client_secret}
    params = {"query": query, "display": 40, "sort": "sim"}
    try:
        resp = requests.get(NAVER_SHOP_URL, headers=headers, params=params, timeout=10)
        if resp.status_code != 200:
            return {"error": f"{resp.status_code} {resp.text[:150]}"}
        items = resp.json().get("items", [])
        prices = [int(it["lprice"]) for it in items if str(it.get("lprice", "")).isdigit() and int(it["lprice"]) > 0]
        if not prices:
            return {"error": f"가격정보 없음 (검색결과 {len(items)}건)"}
        malls = set(it.get("mallName", "") for it in items)
        return {"avg_price": sum(prices) / len(prices), "min_price": min(prices), "mall_count": len(malls)}
    except Exception as e:
        return {"error": str(e)}


def render_discovery_page():
    st.markdown(
        '<div class="dash-header">'
        "<h1>🔍 신상품 발굴</h1>"
        "<p>여러 도매처가 동시에 취급 중인 품목(=지금 제철 신호) 자동 추출 → "
        "월간 검색 트렌드로 2~3개월 뒤 피크 예상 품목 발굴</p>"
        "</div>",
        unsafe_allow_html=True,
    )

    client_id = st.session_state.naver_client_id
    client_secret = st.session_state.naver_client_secret

    with st.expander("⚙️ 네이버 API 키 설정", expanded=not (client_id and client_secret)):
        client_id = st.text_input("Client ID", value=client_id, key="inp_naver_id")
        client_secret = st.text_input("Client Secret", value=client_secret, type="password", key="inp_naver_secret")
        st.session_state.naver_client_id = client_id
        st.session_state.naver_client_secret = client_secret
        st.caption(
            "💡 `.streamlit/secrets.toml`에 `naver_client_id`, `naver_client_secret`을 추가해두면 "
            "새로고침해도 다시 입력 안 해도 됩니다."
        )

    if not client_id or not client_secret:
        st.info("👆 네이버 API 키를 입력하면 분석을 시작할 수 있어요.")
        return

    # ── 분석 대상 자동 추출: 원가비교에서 여러 업체가 동시에 취급 중인 품목 = 지금 제철 신호 ──
    summary = st.session_state._vendor_compare_summary
    if not summary:
        st.warning(
            "⚠️ 아직 분석할 품목이 없습니다. 먼저 **💰 원가비교** 페이지에서 도매처 단가표 엑셀을 업로드해주세요. "
            "거기서 추출된 품목이 여기 분석 대상으로 자동으로 들어옵니다."
        )
        return

    max_vendor_count = max(s["vendor_count"] for s in summary.values())
    st.markdown("#### 🔎 분석 대상 품목 (원가비교에서 자동 추출)")
    min_vendors = st.slider(
        "최소 동시 취급 업체 수 (여러 업체가 동시에 팔고 있을수록 '지금 제철'일 가능성이 높음)",
        min_value=1, max_value=max(max_vendor_count, 1), value=min(2, max_vendor_count),
    )
    candidates = sorted(
        [b for b, s in summary.items() if s["vendor_count"] >= min_vendors],
        key=lambda b: -summary[b]["vendor_count"],
    )
    st.caption(f"📦 {min_vendors}개 이상 업체가 동시 취급 중인 품목: **{len(candidates)}개**")

    MAX_ANALYZE = 30
    selected = st.multiselect(
        f"분석할 품목 선택 (자동으로 채워짐, 최대 {MAX_ANALYZE}개 권장 — 직접 추가/제외 가능)",
        options=sorted(summary.keys()),
        default=candidates[:MAX_ANALYZE],
    )
    if len(selected) > MAX_ANALYZE:
        st.warning(f"⚠️ {MAX_ANALYZE}개 넘으면 API 호출이 많아져서 느려질 수 있어요. 가능하면 줄여주세요.")
    st.caption(
        "⚠️ 네이버 데이터랩은 **같은 요청에 묶인 최대 5개끼리만** 검색량을 직접 비교할 수 있어요. "
        "5개씩 끊어서 조회되며, 묶음번호가 다르면 절대적인 인기도는 직접 비교가 안 됩니다 (묶음 안에서는 정확)."
    )

    target_margin = st.slider("목표 순마진율 구간 (%)", 0, 60, (18, 40))

    with st.expander("⚙️ 채널 수수료율 조정"):
        channel_fees = {}
        for name, default in _CHANNEL_FEES_DEFAULT.items():
            pct = st.number_input(name, 0.0, 50.0, round(default * 100, 2), 0.1, key=f"disc_fee_{name}")
            channel_fees[name] = pct / 100

    if st.button("🔍 월간 트렌드·시장가 분석 시작", type="primary", use_container_width=True):
        keywords = selected
        if not keywords:
            st.error("분석할 품목을 선택해주세요.")
        else:
            with st.spinner(f"{len(keywords)}개 품목 월간 트렌드 분석 중… (네이버 데이터랩 + 쇼핑 조회)"):
                trend_data = fetch_naver_trend_batch(tuple(keywords), client_id, client_secret, time_unit="month")
                today = datetime.date.today()
                current_month = today.month
                rows = []
                for idx, kw in enumerate(keywords):
                    batch_no = idx // 5 + 1  # 5개씩 묶여서 호출되므로, 같은 묶음번호끼리만 인기도 직접비교 가능
                    season = analyze_seasonality(trend_data, kw)
                    shop = fetch_naver_shopping(kw, client_id, client_secret)

                    momentum = season["momentum"] if season["ok"] else None
                    not_declining = (momentum is None) or (momentum > -0.15)

                    if not season["ok"]:
                        fit = f"❓ 판단불가 — {season['debug']}"
                    elif season["is_evergreen"]:
                        fit = "🌱 에버그린형" if not_declining else "🌱 에버그린(하락세)"
                    elif season["peak_month"] is not None:
                        # 2~3개월 뒤(다가오는 시즌)를 내다보고 피크 적합도 판정
                        dist = min(
                            circular_month_distance(season["peak_month"], (current_month + off - 1) % 12 + 1)
                            for off in range(0, 3)
                        )
                        if dist <= 1:
                            fit = "🌸 2~3개월 내 피크 예상" if not_declining else "🌸 피크 예상(하락세)"
                        else:
                            fit = f"⏳ 비시즌(피크 {season['peak_month']}월)"
                    else:
                        fit = "❓ 판단불가"

                    row = {
                        "품목": kw,
                        "동시취급업체수": summary.get(kw, {}).get("vendor_count", "-"),
                        "비교묶음": f"{batch_no}번",
                        "시즌/적합도": fit,
                        "모멘텀%": round(momentum * 100, 1) if momentum is not None else None,
                    }

                    # 원가: 상품마스터에 등록돼 있으면 그 값, 없으면 원가비교에서 찾은 최저원가로 추정
                    cost = None
                    for p in st.session_state.product_master:
                        if kw in p["product_name"] or p["product_name"] in kw:
                            cost = p["cost_price"]
                            break
                    if cost is None:
                        cost = summary.get(kw, {}).get("min_cost")

                    if "error" not in shop:
                        avg = shop["avg_price"]
                        row["시장평균가(원)"] = round(avg)
                        row["판매처수"] = shop["mall_count"]
                        best = None
                        for ch, fee in channel_fees.items():
                            if cost:
                                m = (avg * (1 - fee) - cost) / avg * 100
                                row[f"{ch} 마진%"] = round(m, 1)
                                best = m if best is None else max(best, m)
                            else:
                                row[f"{ch} 마진%"] = "원가없음"
                        lo, hi = target_margin
                        if best is None:
                            row["판정"] = "ℹ️ 원가 확인 불가"
                        elif lo <= best <= hi:
                            row["판정"] = "✅ 목표마진 충족"
                        elif best > hi:
                            row["판정"] = "💎 고마진"
                        elif best < 0:
                            row["판정"] = "🔴 역마진"
                        else:
                            row["판정"] = "⚠️ 마진부족"
                        row["_best"] = best if best is not None else -1000
                    else:
                        row["시장평균가(원)"] = None
                        row["판매처수"] = None
                        row["판정"] = f"⚠️ 조회실패: {shop['error'][:40]}"
                        row["_best"] = -1000

                    row["_log"] = season.get("debug", "")
                    row["_not_declining"] = not_declining
                    rows.append(row)

                st.session_state._discovery_rows = rows

    if st.session_state._discovery_rows:
        st.divider()
        st.markdown("#### 📊 분석 결과")
        rows = st.session_state._discovery_rows
        df = pd.DataFrame(rows).sort_values("_best", ascending=False)
        display_cols = [c for c in df.columns if not c.startswith("_")]
        st.dataframe(df[display_cols].set_index("품목"), use_container_width=True)

        lo, hi = target_margin
        cond_fit = df["시즌/적합도"].isin(["🌸 2~3개월 내 피크 예상", "🌱 에버그린형"])
        cond_margin = df["판정"].isin(["✅ 목표마진 충족", "💎 고마진"])
        recommended = df[cond_fit & cond_margin & df["_not_declining"]]
        st.markdown(f"##### 🏆 추천 품목 (시즌적합·에버그린 + 마진 {lo}~{hi}% 이상)")
        if len(recommended) > 0:
            st.dataframe(recommended[display_cols].set_index("품목"), use_container_width=True)
        else:
            st.info("현재 조건을 충족하는 품목이 없습니다. 품목을 더 추가하거나 마진 구간을 조정해보세요.")

        with st.expander("🔎 품목별 분석 로그 보기"):
            for r in rows:
                st.markdown(f"**{r['품목']}** — {r['_log']}")


# ══════════════════════════════════════════
#  메인 대시보드
# ══════════════════════════════════════════
def render_dashboard():
    # ── 헤더 ──
    st.markdown(
        '<div class="dash-header">'
        "<h1>🛒 쿠팡 주문관리</h1>"
        "<p>발주서 엑셀을 올리면 도매사별로 자동 분류해드립니다</p>"
        "</div>",
        unsafe_allow_html=True,
    )

    # ── 메인 액션: 엑셀(발주서) 업로드 ──
    st.markdown("#### 📁 발주서 엑셀 업로드")
    st.caption(
        "쿠팡 Wing → 주문/배송 → 발주서 조회 → 엑셀 다운로드(DeliveryList_*.xlsx) 파일을 그대로 올려주세요."
    )
    if st.session_state.gsheet_url:
        st.caption("☁️ 구글시트 연동 중 — 업로드한 주문은 자동 저장되어 새로고침해도 유지됩니다.")

    uploaded = st.file_uploader(
        "발주서 엑셀 파일 선택", type=["xlsx"], key="delivery_excel_uploader", label_visibility="collapsed"
    )
    uc1, uc2 = st.columns([3, 1])
    with uc1:
        if uploaded is not None:
            if st.button("📥 업로드한 파일로 주문 불러오기", use_container_width=True, type="primary"):
                try:
                    orders = parse_delivery_excel(uploaded)
                    if not orders:
                        st.warning("⚠️ 파일에서 주문 데이터를 찾지 못했습니다. 파일 양식을 확인해주세요.")
                    else:
                        st.session_state.orders = orders
                        try:
                            save_orders_cache_to_sheet(orders)
                        except Exception:
                            pass
                        st.success(f"✅ {len(orders)}건 주문을 엑셀에서 불러왔습니다!")
                        st.rerun()
                except Exception as e:
                    st.error(f"❌ {e}")
    with uc2:
        if st.button("🗑️ 초기화", use_container_width=True):
            st.session_state.orders = []
            st.rerun()

    # ── 쿠팡 API 자동 동기화 (선택사항, 작게 접어둠) ──
    with st.expander("🔄 쿠팡 API로 자동 동기화 (선택 — IP 화이트리스트 등록 필요)", expanded=False):
        if st.button("API로 오늘 주문 동기화", use_container_width=True):
            ak = st.session_state.api_access_key
            sk = st.session_state.api_secret_key
            vid = st.session_state.api_vendor_id
            if not ak or not sk or not vid:
                st.error("❌ 사이드바 → ⚙️ 쿠팡 API 설정에서 Access Key / Secret Key / Vendor ID를 입력하세요.")
            else:
                with st.spinner("쿠팡 API 연결 중…"):
                    try:
                        orders = fetch_coupang_orders(ak, sk, vid, debug=True)
                        st.session_state.orders = orders
                        try:
                            save_orders_cache_to_sheet(orders)
                        except Exception:
                            pass
                        st.success(f"✅ {len(orders)}건 주문 동기화 완료!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")

    # ── 주문 없음 안내 ──
    if not st.session_state.orders:
        st.markdown(
            "<br><div style='text-align:center;color:#94a3b8;padding:60px 0'>"
            "<div style='font-size:56px'>📭</div>"
            "<h3>수집된 주문이 없습니다</h3>"
            "<p>위에서 발주서 엑셀을 업로드해주세요.</p>"
            "</div>",
            unsafe_allow_html=True,
        )
        return

    # ── VLOOKUP: 주문 ↔ 상품 마스터 매핑 ──
    master_map = {p["option_id"]: p for p in st.session_state.product_master}

    matched   = []
    unmatched = []
    for order in st.session_state.orders:
        master = master_map.get(order["option_id"])
        if master:
            matched.append({**order, "_master": master})
        else:
            unmatched.append(order)

    # ── 전체 요약 메트릭 ──
    total_orders  = len(st.session_state.orders)
    total_revenue = sum(o["sale_price"] * o["quantity"] for o in st.session_state.orders)
    total_profit  = sum(calc_profit(o, o["_master"])["profit"] for o in matched)
    total_fee     = sum(calc_profit(o, o["_master"])["fee"]    for o in matched)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📦 총 주문",   f"{total_orders}건")
    m2.metric("💰 총 매출",   f"₩{total_revenue:,.0f}")
    m3.metric("🏦 수수료 합계", f"₩{total_fee:,.0f}")
    profit_pct = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    m4.metric("📈 예상 순이익", f"₩{total_profit:,.0f}",
              delta=f"{profit_pct:.1f}%",
              delta_color="normal" if total_profit >= 0 else "inverse")

    st.divider()

    # ── 도매사별 그룹핑 ──
    wholesale_groups: dict[str, dict] = {}
    for order in matched:
        name = order["_master"]["wholesale_name"]
        if name not in wholesale_groups:
            wholesale_groups[name] = {"orders": [], "url": order["_master"]["wholesale_url"]}
        wholesale_groups[name]["orders"].append(order)

    # 탭 이름 목록
    tab_labels = [f"🏪 {n}" for n in wholesale_groups]
    if unmatched:
        tab_labels.append("❓ 미매핑 주문")

    if not tab_labels:
        st.warning(
            "⚠️ 주문이 상품 마스터와 매핑되지 않았습니다.\n"
            "사이드바에서 쿠팡등록번호(옵션ID)를 등록해주세요."
        )
        return

    tabs = st.tabs(tab_labels)

    # ── 도매사 탭 렌더링 ──
    wholesale_names = list(wholesale_groups.keys())

    for tab_obj, label in zip(tabs, tab_labels):
        with tab_obj:
            # 미매핑 탭
            if label.startswith("❓"):
                st.warning(f"⚠️ {len(unmatched)}건의 주문이 상품 마스터와 매핑되지 않았습니다.")

                # ── 옵션ID별로 묶어서 한 번에 빠른 등록 (제목/판매가는 주문에서 자동으로 가져옴) ──
                by_option = {}
                for o in unmatched:
                    by_option.setdefault(o["option_id"], []).append(o)

                st.markdown("##### 🆕 옵션ID별 빠른 등록")
                st.caption(
                    "같은 옵션ID는 제목·판매가가 항상 동일하니 자동으로 채워드려요. "
                    "원가(도매가)만 입력하고 저장하면, 해당 옵션ID의 주문 전체가 한 번에 분류됩니다."
                )
                existing_wholesalers = sorted({
                    p["wholesale_name"] for p in st.session_state.product_master if p.get("wholesale_name")
                })

                for opt_id, orders_for_opt in by_option.items():
                    sample = orders_for_opt[0]
                    safe_key = re.sub(r'\W+', '_', opt_id)
                    with st.container(border=True):
                        st.markdown(
                            f"**{sample['product_name']}** "
                            f"&nbsp; <small style='color:#94a3b8'>옵션ID `{opt_id}` · 이 옵션 주문 {len(orders_for_opt)}건</small>",
                            unsafe_allow_html=True,
                        )
                        with st.form(f"quick_unmapped_{safe_key}"):
                            qc1, qc2, qc3 = st.columns(3)
                            q_cost = qc1.number_input("💰 원가(도매가)", min_value=0, value=0, step=500, key=f"um_cost_{safe_key}")
                            q_fee = qc2.number_input("📊 수수료율(%)", min_value=0.0, max_value=100.0, value=10.8, step=0.1, key=f"um_fee_{safe_key}")
                            q_sale = qc3.number_input("🏷️ 판매가(원)", min_value=0, value=int(sample["sale_price"]), step=500, key=f"um_sale_{safe_key}")

                            if existing_wholesalers:
                                q_wname = st.selectbox(
                                    "🏪 도매처", existing_wholesalers + ["+ 새 도매처 직접 입력"], key=f"um_wpick_{safe_key}"
                                )
                                if q_wname == "+ 새 도매처 직접 입력":
                                    q_wname = st.text_input("새 도매처 이름", key=f"um_wnew_{safe_key}")
                            else:
                                q_wname = st.text_input("🏪 도매처", key=f"um_wname_{safe_key}")
                            q_wurl = st.text_input("🔗 도매 주문 페이지 URL", key=f"um_wurl_{safe_key}")

                            if st.form_submit_button("✅ 등록", use_container_width=True, type="primary"):
                                if not q_wname:
                                    st.error("도매처를 입력해주세요.")
                                else:
                                    st.session_state.product_master.append({
                                        "option_id":     opt_id,
                                        "product_name":  sample["product_name"],
                                        "cost_price":    q_cost,
                                        "sale_price":    q_sale,
                                        "fee_rate":      q_fee,
                                        "wholesale_name":q_wname,
                                        "wholesale_url": q_wurl,
                                    })
                                    try:
                                        save_product_master_to_sheet()
                                    except Exception:
                                        pass
                                    st.success(f"✅ 등록 완료! 이 옵션 주문 {len(orders_for_opt)}건이 '{q_wname}' 탭으로 분류됩니다.")
                                    st.rerun()

                st.divider()
                st.markdown("##### 📋 미매핑 주문 상세")
                for o in unmatched:
                    with st.container(border=True):
                        st.markdown(
                            f"**주문 ID:** `{o['order_id']}` | **옵션 ID:** `{o['option_id']}`"
                        )
                        st.markdown(f"**상품명:** {o['product_name']} | 수량: {o['quantity']}개")
                        st.markdown(f"👤 **{o['orderer_name']}** &nbsp; 📞 {o['phone']}", unsafe_allow_html=True)
                        st.markdown(f"📍 {o['address'] or '주소 정보 없음'}")
                        if o["delivery_message"]:
                            st.markdown(f"💬 *{o['delivery_message']}*")

                        quick_copy = (
                            f"{o['product_name']} × {o['quantity']}개 | "
                            f"{o['orderer_name']} {o['phone']} | "
                            f"{o['address']}"
                            + (f" | 요청: {o['delivery_message']}" if o["delivery_message"] else "")
                        )
                        st.code(quick_copy, language=None)
                continue

            # 일반 도매사 탭
            wholesale_name = label.replace("🏪 ", "")
            group = wholesale_groups[wholesale_name]
            grp_orders = group["orders"]
            grp_url    = group["url"]

            grp_revenue = sum(o["sale_price"] * o["quantity"] for o in grp_orders)
            grp_profit  = sum(calc_profit(o, o["_master"])["profit"] for o in grp_orders)
            pct = (grp_profit / grp_revenue * 100) if grp_revenue > 0 else 0
            profit_cls = "profit-pos" if grp_profit >= 0 else "profit-neg"

            # 탭 헤더 + 도매사 바로가기
            hc1, hc2 = st.columns([3, 1])
            with hc1:
                st.markdown(
                    f'<div class="tab-header">'
                    f"<h4>🏪 {wholesale_name}</h4>"
                    f"<span>주문 {len(grp_orders)}건 &nbsp;|&nbsp; "
                    f"매출 ₩{grp_revenue:,.0f} &nbsp;|&nbsp; "
                    f"예상순이익 <span class='{profit_cls}'>₩{grp_profit:,.0f} ({pct:.1f}%)</span>"
                    f"</span></div>",
                    unsafe_allow_html=True,
                )
            with hc2:
                if grp_url:
                    st.link_button(
                        "🚀 도매 사이트 바로가기",
                        grp_url,
                        use_container_width=True,
                        type="primary",
                    )
                else:
                    st.button("🔗 URL 미등록", disabled=True,
                              use_container_width=True, key=f"no_url_{wholesale_name}")

            # ── 주문 카드 ──
            for o in grp_orders:
                master = o["_master"]
                pf = calc_profit(o, master)

                with st.container(border=True):
                    c1, c2, c3 = st.columns([3, 3, 1])

                    with c1:
                        st.markdown(f"**{master['product_name']}**")
                        st.caption(f"주문 ID: `{o['order_id']}` | 옵션 ID: `{o['option_id']}`")
                        st.caption(f"🕒 {o['ordered_at']}")
                        st.markdown(f"👤 **{o['orderer_name']}** &nbsp; 📞 {o['phone']}", unsafe_allow_html=True)
                        st.markdown(f"📍 {o['address'] or '주소 정보 없음'}")
                        if o["delivery_message"]:
                            st.markdown(f"💬 *{o['delivery_message']}*")

                    with c2:
                        mc1, mc2, mc3, mc4 = st.columns(4)
                        mc1.metric("수량",   f"{o['quantity']}개")
                        mc2.metric("매출",   f"₩{pf['revenue']:,.0f}")
                        mc3.metric("원가",   f"₩{pf['cost']:,.0f}")
                        delta_color = "normal" if pf["profit"] >= 0 else "inverse"
                        mc4.metric(
                            "순이익",
                            f"₩{pf['profit']:,.0f}",
                            delta=f"{pf['margin']:.1f}%",
                            delta_color=delta_color,
                        )

                    with c3:
                        st.write("")   # 수직 정렬 보정
                        st.write("")
                        dialog_key = f"dlg_{o['order_id']}_{wholesale_name}"
                        if st.button(
                            "📋 발주 정보\n새창 열기",
                            key=dialog_key,
                            use_container_width=True,
                        ):
                            order_dialog(o, master)

                    # 도매사이트 발주 입력칸에 바로 붙여넣기용 (클릭 한 번 없이 카드에서 바로 복사)
                    quick_copy = (
                        f"{master['product_name']} × {o['quantity']}개 | "
                        f"{o['orderer_name']} {o['phone']} | "
                        f"{o['address']}"
                        + (f" | 요청: {o['delivery_message']}" if o["delivery_message"] else "")
                    )
                    st.code(quick_copy, language=None)


# ══════════════════════════════════════════
#  신상품 발굴 페이지 — 도매처별 단가표 업로드 → 동일품목 비교 → 빠른 등록
# ══════════════════════════════════════════
def render_sourcing_page():
    st.markdown(
        '<div class="dash-header">'
        "<h1>💰 원가비교</h1>"
        "<p>도매처별 단가표 업로드 → 품목명은 같게, 중량·등급·개수구간은 정확히 분리해서 비교</p>"
        "</div>",
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader(
        "도매처별 단가표 엑셀을 업로드하세요 (도매처마다 한 파일씩, 여러 개 동시 업로드 가능)",
        type=["xlsx"], accept_multiple_files=True, key="vendor_price_files",
    )

    if not uploaded_files:
        cached = st.session_state.get("_vendor_items_cache") or []
        if not cached and st.session_state.gsheet_url:
            try:
                cached = load_vendor_items_from_sheet()
                if cached:
                    st.session_state._vendor_items_cache = cached
            except Exception:
                pass
        if cached:
            all_items = cached
            cc1, cc2 = st.columns([4, 1])
            cc1.success(f"📦 이전에 업로드한 데이터를 사용 중입니다 (총 {len(all_items)}개 항목). 새 파일을 올리면 자동으로 갱신돼요.")
            if cc2.button("🗑️ 초기화", use_container_width=True):
                st.session_state._vendor_items_cache = []
                try:
                    save_vendor_items_to_sheet([])
                except Exception:
                    pass
                st.rerun()
        else:
            st.info(
                "💡 '사과', '시나노사과'처럼 다른 품종은 따로 유지하고, 같은 품목이어도 "
                "**1kg/2kg, 대과/중과/소과처럼 규격이 다르면 별도 행으로 분리해서** 도매처별 원가를 비교해드립니다."
            )
            return
    else:
        def _guess(colnames, keys):
            for c in colnames:
                if any(k in str(c) for k in keys):
                    return c
            return colnames[0]

        all_items = []
        for f in uploaded_files:
            df_preview = pd.read_excel(f)
            cols = list(df_preview.columns)

            vendor_name = f.name.rsplit(".", 1)[0]
            name_col = _guess(cols, ["상품명"])
            opt_match = next((c for c in cols if "옵션명" in str(c)), None)
            option_col = opt_match if opt_match else "(없음)"
            price_col = _guess(cols, ["공급가", "원가"])

            with st.expander(f"📄 {f.name} — 컬럼 매칭 확인 (자동 인식됨, 틀렸을 때만 펼쳐서 수정)", expanded=False):
                st.dataframe(df_preview.head(3), use_container_width=True)

                vendor_name = st.text_input(
                    "이 파일의 도매처 이름", value=vendor_name, key=f"sc_vendor_{f.name}"
                )
                cc1, cc2, cc3 = st.columns(3)
                name_col = cc1.selectbox(
                    "품목명 컬럼", cols, index=cols.index(name_col), key=f"sc_name_{f.name}"
                )
                opt_options = ["(없음)"] + cols
                option_col = cc2.selectbox(
                    "옵션명 컬럼 (있으면 — 품목명과 합쳐서 분석)", opt_options,
                    index=opt_options.index(option_col),
                    key=f"sc_opt_{f.name}",
                )
                price_col = cc3.selectbox(
                    "원가(공급가) 컬럼", cols, index=cols.index(price_col), key=f"sc_price_{f.name}"
                )

            f.seek(0)
            items = parse_vendor_excel(f, vendor_name, name_col, option_col, price_col)
            all_items.extend(items)
            st.caption(f"✅ {f.name}: {len(items)}개 항목 인식 (도매처: {vendor_name})")

        # 페이지를 이동했다 돌아와도, 새로고침해도, 세션이 끊겨도 다시 업로드 안 해도 되게 저장
        st.session_state._vendor_items_cache = all_items
        try:
            save_vendor_items_to_sheet(all_items)
        except Exception:
            pass  # 구글시트 미설정 시 조용히 무시 (세션 내에서는 계속 사용 가능)

    if not all_items:
        return

    st.divider()
    tree = group_vendor_items(all_items)  # { 품목명: { 규격라벨: {도매처: 원가, "_원본예시": str} } }
    total_specs = sum(len(specs) for specs in tree.values())
    st.success(f"총 {len(all_items)}개 항목 → **{len(tree)}개 품목 · {total_specs}개 규격(중량×등급)**으로 정리되었습니다.")

    # 신상품 발굴 페이지에서 쓸 수 있도록, 품목별 '동시 취급 업체 수' 요약을 세션에 저장
    # (여러 업체가 동시에 취급 중 = 지금 시중에 풀린 제철 신호로 간주)
    summary = {}
    for base, specs in tree.items():
        vendors = set()
        all_costs = []
        for spec_vendors in specs.values():
            for v, c in spec_vendors.items():
                if v == "_원본예시":
                    continue
                vendors.add(v)
                all_costs.append(c)
        summary[base] = {
            "vendor_count": len(vendors),
            "vendors": sorted(vendors),
            "min_cost": min(all_costs) if all_costs else None,
        }
    st.session_state._vendor_compare_summary = summary

    # ── 품목 검색 (검색하기 전에는 비교 카드를 그리지 않음 — 미리 다 그리면 느려짐) ──
    st.markdown("#### 🔎 품목 검색")
    search = st.text_input(
        "품목명을 입력하면 그 품목의 모든 규격을 도매처별로 비교해서 보여드립니다",
        "", placeholder="예: 시나노사과 (입력 후 Enter)",
        label_visibility="collapsed",
    )

    if not search.strip():
        st.info("👆 검색어를 입력하면 그 품목의 도매처별 규격 비교표가 여기 표시됩니다.")
        return

    base_names = [b for b in tree if search.strip() in b]
    st.caption(f"'{search.strip()}' 검색결과 {len(base_names)}개 품목")

    if not base_names:
        st.warning("검색 결과가 없습니다. 다른 키워드로 시도해보세요.")
        return

    MAX_RESULTS = 20
    if len(base_names) > MAX_RESULTS:
        st.caption(f"⚠️ 결과가 많아 상위 {MAX_RESULTS}개만 표시합니다. 더 구체적인 키워드로 검색해보세요.")
        base_names = base_names[:MAX_RESULTS]

    def _highlight_cheapest(row, vendor_cols):
        """행에서 최저가 셀은 초록, 최고가 셀은 옅은 빨강으로 표시."""
        styles = [''] * len(row)
        vals = pd.to_numeric(row[vendor_cols], errors='coerce')
        valid = vals.dropna()
        if len(valid) < 1:
            return styles
        min_v, max_v = valid.min(), valid.max()
        for i, col in enumerate(row.index):
            if col not in vendor_cols or pd.isna(row[col]):
                continue
            if row[col] == min_v and min_v != max_v:
                styles[i] = 'background-color:#bbf7d0; color:#166534; font-weight:700'
            elif row[col] == max_v and min_v != max_v:
                styles[i] = 'background-color:#fecaca; color:#991b1b'
        return styles

    for base in base_names:
        specs = tree[base]
        with st.container(border=True):
            st.markdown(
                f"**🍎 {base}** "
                f"&nbsp; <small style='color:#94a3b8'>규격 {len(specs)}종류</small>",
                unsafe_allow_html=True,
            )

            # 규격(행) × 도매처(열) 비교 매트릭스 — 최저가는 초록, 최고가는 빨강으로 강조
            vendor_set = sorted({v for spec_vendors in specs.values() for v in spec_vendors if v != "_원본예시"})
            table_rows = []
            for spec_label, vendor_costs in specs.items():
                row = {"규격": spec_label}
                for v in vendor_set:
                    row[v] = vendor_costs.get(v)  # 숫자 그대로 (None 허용)
                table_rows.append(row)

            spec_df = pd.DataFrame(table_rows).set_index("규격")
            styled = (
                spec_df.style
                .apply(_highlight_cheapest, vendor_cols=vendor_set, axis=1)
                .format(lambda x: f"₩{x:,.0f}" if pd.notna(x) else "-")
            )
            st.dataframe(styled, use_container_width=True)
            st.caption("🟢 최저가 · 🔴 최고가")

            with st.expander(f"📥 '{base}' 특정 규격 상품마스터에 빠른 등록"):
                spec_options = list(specs.keys())
                picked_spec = st.selectbox("등록할 규격 선택", spec_options, key=f"pick_spec_{re.sub(r'\W+', '_', base)}")
                vendor_costs = {v: c for v, c in specs[picked_spec].items() if v != "_원본예시"}
                cheapest_vendor = min(vendor_costs, key=vendor_costs.get)
                cheapest_cost = vendor_costs[cheapest_vendor]
                example_name = specs[picked_spec].get("_원본예시", f"{base} {picked_spec}")
                safe_key = re.sub(r'\W+', '_', f"{base}_{picked_spec}")

                with st.form(f"quickreg_{safe_key}"):
                    q_oid = st.text_input(
                        "🔑 쿠팡등록번호(옵션ID) — 아직 쿠팡에 등록 전이면 비워두세요 (나중에 사이드바에서 수정 가능)",
                        key=f"q_oid_{safe_key}",
                    )
                    q_pname = st.text_input("📦 제품이름", value=f"{base} {picked_spec}".strip(), key=f"q_pname_{safe_key}")
                    q_cost  = st.number_input("💰 원가(도매가)", min_value=0, value=int(cheapest_cost), step=500, key=f"q_cost_{safe_key}")
                    q_sale  = st.number_input("🏷️ 쿠팡 판매가(원)", min_value=0, value=int(cheapest_cost * 1.6), step=500, key=f"q_sale_{safe_key}")
                    q_fee   = st.number_input("📊 판매 수수료율(%)", min_value=0.0, max_value=100.0, value=10.8, step=0.1, key=f"q_fee_{safe_key}")
                    q_wname = st.selectbox("🏪 도매처 (최저가순)", sorted(vendor_costs, key=vendor_costs.get), key=f"q_wname_{safe_key}")
                    q_wurl  = st.text_input("🔗 도매 주문 페이지 URL", key=f"q_wurl_{safe_key}")
                    st.caption(f"원본 표기 예시: {example_name}")

                    if st.form_submit_button("✅ 상품마스터에 등록", use_container_width=True, type="primary"):
                        new_oid = q_oid.strip() or f"PENDING-{safe_key}"
                        if new_oid in [p["option_id"] for p in st.session_state.product_master]:
                            st.warning("이미 등록된 옵션ID입니다.")
                        else:
                            st.session_state.product_master.append({
                                "option_id":     new_oid,
                                "product_name":  q_pname,
                                "cost_price":    q_cost,
                                "sale_price":    q_sale,
                                "fee_rate":      q_fee,
                                "wholesale_name":q_wname,
                                "wholesale_url": q_wurl,
                            })
                            try:
                                save_product_master_to_sheet()
                            except Exception:
                                pass
                            if not q_oid.strip():
                                st.success(
                                    f"✅ '{q_pname}' 등록 완료! 옵션ID는 임시값(PENDING-...)입니다. "
                                    "쿠팡에 실제 등록 후 사이드바 → 🍊 상품 마스터에서 ✏️ 눌러 진짜 옵션ID로 수정해주세요."
                                )
                            else:
                                st.success(f"✅ '{q_pname}' 등록 완료!")
                            st.rerun()


# ══════════════════════════════════════════
#  진입점
# ══════════════════════════════════════════
def main():
    # F5 새로고침해도 로그인 유지: URL에 유효한 토큰이 있으면 자동 로그인 처리
    if not st.session_state.logged_in:
        _check_url_auth()

    # 로그인 상태가 아니면 로그인 페이지만 표시
    if not st.session_state.logged_in:
        login_page()
        return

    # 구글시트가 연동되어 있으면, 세션당 1회 자동으로 상품마스터/마지막 주문/원가비교 데이터를 불러옴
    # (새로고침/재접속할 때마다 다시 입력·업로드할 필요 없게)
    if st.session_state.gsheet_url and not st.session_state._synced_from_sheet:
        try:
            if not st.session_state.product_master:
                st.session_state.product_master = load_product_master_from_sheet()
            if not st.session_state.orders:
                st.session_state.orders = load_orders_cache_from_sheet()
            if not st.session_state._vendor_items_cache:
                st.session_state._vendor_items_cache = load_vendor_items_from_sheet()
            if st.session_state._vendor_items_cache and not st.session_state._vendor_compare_summary:
                tree = group_vendor_items(st.session_state._vendor_items_cache)
                summary = {}
                for base, specs in tree.items():
                    vendors, all_costs = set(), []
                    for spec_vendors in specs.values():
                        for v, c in spec_vendors.items():
                            if v == "_원본예시":
                                continue
                            vendors.add(v)
                            all_costs.append(c)
                    summary[base] = {
                        "vendor_count": len(vendors), "vendors": sorted(vendors),
                        "min_cost": min(all_costs) if all_costs else None,
                    }
                st.session_state._vendor_compare_summary = summary
        except Exception:
            pass  # 연동 설정이 미완료여도 앱은 정상 동작
        st.session_state._synced_from_sheet = True

    # 사이드바는 공통으로 항상 표시
    render_sidebar()

    # 상단 3버튼 네비게이션 (페이지 전환 + 주요 액션을 겸함)
    if "current_page" not in st.session_state:
        st.session_state.current_page = "order"

    nc1, nc2, nc3 = st.columns(3)
    if nc1.button(
        "🛒 쿠팡 주문관리", use_container_width=True,
        type="primary" if st.session_state.current_page == "order" else "secondary",
    ):
        st.session_state.current_page = "order"
        st.rerun()
    if nc2.button(
        "💰 원가비교", use_container_width=True,
        type="primary" if st.session_state.current_page == "compare" else "secondary",
    ):
        st.session_state.current_page = "compare"
        st.rerun()
    if nc3.button(
        "🔍 신상품 발굴", use_container_width=True,
        type="primary" if st.session_state.current_page == "discovery" else "secondary",
    ):
        st.session_state.current_page = "discovery"
        st.rerun()

    st.divider()

    if st.session_state.current_page == "order":
        render_dashboard()
    elif st.session_state.current_page == "compare":
        render_sourcing_page()
    else:
        render_discovery_page()


if __name__ == "__main__":
    main()
